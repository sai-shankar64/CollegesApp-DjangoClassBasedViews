import os,sys,django
os.environ['DJANGO_SETTINGS_MODULE'] = os.getcwd().split("\\")[-1]+'.settings'
django.setup()
import click
import MySQLdb
from openpyxl import load_workbook
from onlineapp.models import *

@click.group()
def cli():
    """Commands for db import/export"""
    pass

@cli.command()
def createdb():
    """Creates DataBase"""
    db=MySQLdb.connect(host="localhost",user="root",passwd="shankar123")
    cursor=db.cursor()
    sql="CREATE DATABASE IF NOT EXISTS onlinedb"
    cursor.execute(sql)
    cursor.execute("USE onlinedb")
    sql="CREATE TABLE IF NOT EXISTS COLLEGE (NAME VARCHAR(128),LOCATION VARCHAR(64),ACRONYM VARCHAR(10),CONTACT VARCHAR(50),PRIMARY KEY(ACRONYM))"
    cursor.execute(sql)
    sql = "CREATE TABLE IF NOT EXISTS STUDENT (NAME VARCHAR(128),TRANSFORM INT,FROM_CUSTOM_BASE26 INT,GET_PIG_LATIN INT,TOP_CHARS INT,TOTAL INT,PRIMARY KEY(STUDENT),FOREIGN KEY (STUDENT) REFERENCES STUDENT(MARKSID))"
    cursor.execute(sql)
    print("DataBase created with tables")
    db.commit()
    db.close()

@cli.command()
def populatedb():
    """Insert data to database"""
    colg_wb=load_workbook(filename="students.xlsx")
    colg_ws=colg_wb["Colleges"]
    rows=colg_ws.rows
    rows.__next__()
    for row in rows:
        values=[]
        for cell in row:
            values.append(cell.value)
        colg=College(name=values[0],location=values[2],acronym=values[1],contact=values[3])
        colg.save()

    stud_wb=load_workbook(filename="students.xlsx")
    stud_ws=stud_wb["Current"]

    rows=stud_ws.rows
    rows.__next__()
    for row in rows:
        values=[]
        for cell in row:
            values.append(cell.value)
        student=Student(name=values[0],email=values[2],db_folder=values[3],college = College.objects.get(acronym=values[1]))
        student.save()

    stud_ws = stud_wb["Deletions"]
    rows = stud_ws.rows
    rows.__next__()
    for row in rows:
        values = []
        for cell in row:
            values.append(cell.value)
        student = Student(name=values[0], email=values[2], db_folder=values[3],
                          college=College.objects.get(acronym=values[1]),dropped_out=True)
        student.save()


    marks_wb=load_workbook(filename="studentmarks.xlsx")
    marks_ws=marks_wb["MockResults2016"]
    rows=marks_ws.rows
    rows.__next__()
    for row in rows:
        print(row)
        values=[]
        for cell in row:
            values.append(cell.value)
        dbname=values[0]
        splittedvalues=dbname.split('_')
        dbname=splittedvalues[2]
        marks=MockTest1(problem1=values[1],problem2=values[2],problem3=values[3],problem4=values[4],total=values[5],student=Student.objects.get(db_folder=dbname))
        marks.save()

@cli.command()
def cleardata():
    """Cleares table in the databases"""
    pass

@cli.command()
def dropdb():
    """Drop database"""
    pass

if __name__=="__main__":
    cli()